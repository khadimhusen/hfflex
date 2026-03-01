from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from .filters import StockFilter
from .models import Stockdetail


def stocklist(request):
    Stock_list = Stockdetail.objects.all()

    myFilter = StockFilter(request.GET, Stock_list)
    stock_queryset = myFilter.qs
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-stock.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'), )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Stock'
    columns = ['Material', 'Type', 'Grade', 'Size', 'Micron', 'rate', 'qc_status',
               'recieved', 'used',  'balance', 'available', 'alloted', 'created']
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    for mate in stock_queryset:
        row_num += 1
        mateitem=""
        if mate.content_type.model=="prodreport":
            mateitem=str(mate.prodreports.all().first().prodprocess.job)

        row = [str(mate.materialname), str(mate.item_mat_type), str(mate.item_grade),
               mate.size, mate.micron, mate.rate, mate.qc_status, mate.recieved, mate.used,
               mate.balance, mate.available, mate.alloted,
               mate.created, mateitem]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
