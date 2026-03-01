from datetime import datetime

from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from openpyxl import Workbook
from django.http import HttpResponse
from .filters import PoFilter, PoItemFilter
from .models import Po, PoItem


def xlpolist(request):
    q = request.GET.get('q', None)
    if q != None:
        po_list = Po.objects.filter(state=q)
    else:
        po_list = Po.objects.all()

    myFilter = PoFilter(request.GET, po_list)
    po_list = myFilter.qs.select_related('supplier', 'createdby', 'approvedby')

    page = request.GET.get('page', 1)
    paginator = Paginator(po_list, 100)
    try:
        pos = paginator.page(page)
    except PageNotAnInteger:
        pos = paginator.page(1)
    except EmptyPage:
        pos = paginator.page(paginator.num_pages)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-job.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'), )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'PO'
    columns = ['id', 'Supplier', 'Createdby', 'Approvedby', 'Approved Date', 'Status']
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    for po in pos:
        row_num += 1
        row = [po.id, str(po.supplier), str(po.createdby or ""), str(po.approvedby or ""),
               po.approve_date.strftime('%d-%m-%Y') if po.approve_date else '', str(po.status or "")]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response


def xlpoitemlist(request):
    q = request.GET.get('q', None)
    if q != None:
        poitem_list = PoItem.objects.filter(state=q)
    else:
        poitem_list = PoItem.objects.all()

    myFilter = PoItemFilter(request.GET, poitem_list)
    poitem_list = myFilter.qs.select_related('purchaseorder__supplier', 'createdby')

    page = request.GET.get('page', 1)
    paginator = Paginator(poitem_list, 100)
    try:
        pos = paginator.page(page)
    except PageNotAnInteger:
        pos = paginator.page(1)
    except EmptyPage:
        pos = paginator.page(paginator.num_pages)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-job.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'), )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'POITEM'
    #	Po	Desc.	Qty	Rate	Amount	Rec Qty	Created
    columns = ['id', 'PURCHASE ORDER', 'Description', 'Unit', 'Rate', 'Ord. Qty', 'Rec. Qty','Pending qty', 'Status', 'Approval Date']
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    for po in pos:
        row_num += 1
        row = [po.id, str(po.purchaseorder), po.description,
               str(po.unit), po.rate, po.qty, po.rec_qty,(po.qty or 0 - po.rec_qty or 0),
               po.purchaseorder.status,po.purchaseorder.approve_date ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
