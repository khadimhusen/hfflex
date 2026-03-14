from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from .filters import JobProcessFilter, JobFilter, JobMaterialFilter
from .models import JobProcess, Job, JobMaterial
from openpyxl.styles import Alignment
from employee.models import Department


def joblist(request):
    q = request.GET.get('q', None)
    if q is not None:
        job_list = Job.objects.filter(jobstatus=q)
    else:
        job_list = Job.objects.all()

    myFilter = JobFilter(request.GET, job_list)
    job_list = myFilter.qs.select_related(
        'itemmaster', 'unit', 'joborder'
    )

    # check director ONCE before the loop
    is_director = Department.objects.filter(
        department_name="directors", user=request.user
    ).exists()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-job.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Job'

    if is_director:
        columns = ['id', 'Itemmaster', 'Quantity', 'Unit', 'Kg', 'Status', 'Size', 'Waste%', 'Margin', 'per kg']
    else:
        columns = ['id', 'Itemmaster', 'Quantity', 'Unit', 'Kg', 'Status', 'Size', 'Waste%']

    worksheet.column_dimensions['B'].width = 20

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=row_num, column=col_num).value = column_title

    if is_director:
        # pre-fetch all job ids in this queryset
        job_ids = list(job_list.values_list('id', flat=True))

        # pull all production summary data in bulk — one set of queries total
        from production.models import ProdInput, ProdReport, Stockdetail
        from django.contrib.contenttypes.models import ContentType
        from django.db.models import Sum

        report_qs = ProdReport.objects.filter(
            prodprocess__job_id__in=job_ids
        ).values('id', 'prodprocess__job_id')

        # map: job_id -> [report_ids]
        job_report_map = {}
        for r in report_qs:
            job_report_map.setdefault(r['prodprocess__job_id'], []).append(r['id'])

        all_report_ids = [rid for rids in job_report_map.values() for rid in rids]

        # bulk fetch all inputs
        inputs = ProdInput.objects.filter(
            prodreport__in=all_report_ids
        ).select_related(
            'material__materialname',
            'material__item_mat_type',
            'material__item_grade',
            'prodreport__prodprocess',
        ).values(
            'prodreport__prodprocess__job_id',
            'material__rate',
            'inputqty',
            'wtgain',
            'material__materialname__name',
            'material__item_mat_type__mat_type',
            'material__item_grade__grade',
            'material__size',
            'material__micron',
        )

        # bulk fetch all outputs
        ct = ContentType.objects.get_for_model(ProdReport)
        outputs = Stockdetail.objects.filter(
            content_type=ct,
            object_id__in=all_report_ids,
        ).select_related(
            'materialname', 'item_mat_type', 'item_grade'
        ).extra(
            select={'job_id': '''
                SELECT prodprocess__job_id FROM production_prodreport
                WHERE production_prodreport.id = production_stockdetail.object_id
            '''}
        )

        # simpler: get job_id via report map (reverse lookup)
        report_to_job = {}
        for job_id, rids in job_report_map.items():
            for rid in rids:
                report_to_job[rid] = job_id

        # build per-job inputdetail dict
        from collections import defaultdict
        job_inputdetail = defaultdict(dict)

        for inp in inputs:
            jid  = inp['prodreport__prodprocess__job_id']
            key  = (f"{inp['material__materialname__name']} - "
                    f"{inp['material__item_mat_type__mat_type']} - "
                    f"{inp['material__item_grade__grade']}-"
                    f"{inp['material__size']}mm X {inp['material__micron']}mic")
            qty  = round(-(inp['wtgain'] or 0), 3)
            amt  = round((inp['material__rate'] or 0) * (inp['inputqty'] or 0), 3)
            if key not in job_inputdetail[jid]:
                job_inputdetail[jid][key] = {"qty": qty, "amount": amt}
            else:
                job_inputdetail[jid][key]["qty"]    += qty
                job_inputdetail[jid][key]["amount"] += amt

        out_qs = Stockdetail.objects.filter(
            content_type=ct,
            object_id__in=all_report_ids,
        ).values('object_id', 'recieved',
                 'materialname__name',
                 'item_mat_type__mat_type',
                 'item_grade__grade',
                 'size', 'micron')

        for out in out_qs:
            jid = report_to_job.get(out['object_id'])
            if jid is None:
                continue
            key = (f"{out['materialname__name']} - "
                   f"{out['item_mat_type__mat_type']} - "
                   f"{out['item_grade__grade']}-"
                   f"{out['size']}mm X {out['micron']}mic")
            qty = round((out['recieved'] or 0), 3)
            if key not in job_inputdetail[jid]:
                job_inputdetail[jid][key] = {"qty": qty, "amount": 0}
            else:
                job_inputdetail[jid][key]["qty"] += qty

        # compute cost, netoutput per job from inputdetail
        job_cost      = {}
        job_netoutput = {}
        for jid, detail in job_inputdetail.items():
            cost      = 0
            netoutput = 0
            for key, val in detail.items():
                if val["qty"] < -0.0001:
                    cost += val["amount"]
                elif val["qty"] > 0.0001 and 'WASTE' not in key:
                    netoutput = round(netoutput + val["qty"], 3)
            job_cost[jid]      = round(cost, 3)
            job_netoutput[jid] = netoutput

    # also pre-compute jobwaste in bulk the same way (avoid per-job DB hit)
    # jobwaste is already stored as job.job_waste (saved field) — use that instead
    for job in job_list:
        row_num += 1
        if is_director:
            jid      = job.id
            cost     = job_cost.get(jid, 0)
            netout   = job_netoutput.get(jid, 0)
            salecost = round(netout * job.kgrate, 3)
            diff     = round(salecost - cost, 3)
            per_kg   = round(diff / netout, 3) if netout else 0
            row = [
                job.id, str(job.itemname), job.quantity, str(job.unit),
                job.kgqty, str(job.jobstatus), job.film_size,
                job.job_waste,   # ← stored field, no DB hit
                diff, per_kg,
            ]
        else:
            row = [
                job.id, str(job.itemname), job.quantity, str(job.unit),
                job.kgqty, str(job.jobstatus), job.film_size,
                job.job_waste,   # ← stored field, no DB hit
            ]

        for col_num, cell_value in enumerate(row, 1):
            worksheet.cell(row=row_num, column=col_num).value = cell_value

    workbook.save(response)
    return response


def processlist(request, status=None):

    if status == 'All':
        proc_list = JobProcess.objects.all().order_by('-job__film_size')
    else:
        proc_list = JobProcess.objects.filter(process__process=status).order_by('-job__film_size')
    myFilter = JobProcessFilter(request.GET, proc_list)
    process_queryset = myFilter.qs
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-{status}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),status=status )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = str(status)

    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # Apply to column dimensions
    worksheet.column_dimensions['A'].alignment = wrap_alignment

    columns = ['Job', 'Process', 'Qty', 'Unit','Pouch Type', 'Pouch Qty', 'Status', 'Size','Cylinder']
    col_width = {"A": 100, "B": 10, "C": 10, "D": 10, "E": 30, "F": 10, "G": 10, "H":10,"I":10}
    for col in col_width:
        worksheet.column_dimensions[col].width = col_width[col]


    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for process in process_queryset:
        row_num += 1
        if process.process.process =="Lamination":
            ply= " " + str(process.process_count + 1)+ "Ply"
        else:
            ply=""

        row = [str(process.job) + ply, str(process.process), process.qty, str(process.unit), str(process.job.pouch_type),process.job.totalpouch,
               str(process.status), process.job.film_size, process.job.itemmaster.cylinder_status]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrap_alignment

    workbook.save(response)
    return response


def jobmaterialexcel(request):


    q = request.GET.get('q', None)
    if q != None:
        print(q)
        jobmaterial_list = JobMaterial.objects.select_related('materialname', 'item_mat_type',
                                                              'item_grade').filter(jobstatus=q)
    else:
        print(q)
        jobmaterial_list = JobMaterial.objects.select_related('materialname', 'item_mat_type',
                                                              'item_grade').all()

    myFilter = JobMaterialFilter(request.GET, jobmaterial_list)
    jobmaterial_list = myFilter.qs

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-job.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'), )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Requried Material'
    columns = ['Job', 'status', 'Material', 'Type', 'Grade', 'Size', 'Micron', 'Required',
               'Available', 'To Order', 'Ordered','Rec','opt1', 'opt2']
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    for mate in jobmaterial_list:
        row_num += 1
        row = [str(mate.job), str(mate.job.jobstatus), str(mate.materialname), str(mate.item_mat_type),
               str(mate.item_grade), mate.size, mate.micron, mate.required, mate.avail,
               mate.to_order, mate.orderedqty, mate.receivedqty]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
