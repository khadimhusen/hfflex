from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from .filters import ItemmasterFilter
from .models import ItemMaster


def itemlistexcel(request):

    q = request.GET.get('q', None)
    if q != None:
        item_list = ItemMaster.objects.filter(jobstatus=q)
    else:
        item_list = ItemMaster.objects.all()

    myFilter = ItemmasterFilter(request.GET, item_list)
    item_list = myFilter.qs
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-itemmaster.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'), )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Job'
    columns = ['id', 'Name', 'customer', 'cyl_len', 'circum', 'rep_len', 'openwidth', 'ups', 'repeat']

    worksheet.column_dimensions['B'].width = 20
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    for item in item_list:
        row_num += 1
        row = [item.id, item.itemname, str(item.itemcustomer), item.cyl_length,item.cyl_circum, item.replength, item.openwidth,
               item.no_of_ups, item.no_of_repeat]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
